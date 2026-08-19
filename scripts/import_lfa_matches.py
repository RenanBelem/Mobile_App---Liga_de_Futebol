"""
Importa TODAS as partidas de TODAS as temporadas a partir de fonte/dados/LFA.xlsx
e regenera fonte/dados/json/partidas.json.

Uso:
    python scripts/import_lfa_matches.py            # gera o arquivo final
    python scripts/import_lfa_matches.py --dry-run   # só mostra o resumo, não escreve nada

Estrutura da planilha (por aba "APERTURA/CLAUSURA <ano>"):
  - Colunas A-D: Time mandante | Placar mandante | Placar visitante | Time visitante
  - Coluna E: rótulo da rodada/fase (mesclado sobre o bloco de partidas daquela rodada),
    ex.: "1º DATA CECILIA", "3º LUGAR CANTONA", "SEMI WLADIMIR"...
  - Coluna F: rótulo da fase maior (mesclado sobre várias rodadas),
    ex.: "1º FASE (FASE DE LIGA...)", "PLAYOFFS", "1º FASE SÉRIE B...", "1º FASE FEMININO..."
  - Blocos de partidas são separados por linhas em branco.

O nome da competição (Cecília/Caszely/Cantona/Wladimir/Sissi/Foice/Recopa) é decidido assim:
  1. Se o rótulo da rodada (coluna E) menciona explicitamente a competição, usa-a.
  2. Caso contrário, usa o contexto da fase maior (coluna F): "SÉRIE B" -> Wladimir,
     "FEMININO" -> Sissi, senão -> Cecília (competição principal/liga).

Como a planilha não possui datas de calendário, cada bloco de partidas recebe uma data
sintética sequencial (início da competição + 7 dias por bloco), preservando a ordem
cronológica relativa dos jogos. Ajustes finos de data podem ser feitos depois pela tela
de administração.
"""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import date, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "fonte" / "dados" / "LFA.xlsx"
JSON_DIR = ROOT / "fonte" / "dados" / "json"
OUT_PATH = JSON_DIR / "partidas.json"

SHEETS_TO_SKIP = {"CLAUSURA 2026"}  # temporada ainda sem nenhuma partida cadastrada

SHEET_TO_SEASON = {
    "APERTURA 2022": "season-2022-a",
    "CLAUSURA 2022": "season-2022-c",
    "APERTURA 2023": "season-2023-a",
    "CLAUSURA 2023": "season-2023-c",
    "APERTURA 2024": "season-2024-a",
    "CLAUSURA 2024": "season-2024-c",
    "APERTURA 2025": "season-2025-a",
    "CLAUSURA 2025": "season-2025-c",
    "APERTURA 2026": "season-2026-a",
}

# Ordem de checagem importa: a primeira competição encontrada na string é usada.
COMPETITION_TOKENS = [
    ("CECILIA", "taca-cecilia"),
    ("CASZELY", "copa-carlos-caszely"),
    ("CANTONA", "copa-eric-cantona"),
    ("WLADIMIR", "taca-wladimir-rodrigues"),
    ("SISSI", "taca-sissi"),
    ("FOICE", "copa-foice"),
    ("RECOPA", "recopa"),
]

# Contexto de fase (coluna F) -> slug de competição default quando a rodada (coluna E)
# não menciona a competição explicitamente.
PHASE_CONTEXT_DEFAULT = [
    ("SERIE B", "taca-wladimir-rodrigues"),
    ("FEMININO", "taca-sissi"),
]
DEFAULT_COMPETITION_SLUG = "taca-cecilia"

TEAM_ALIASES: dict[str, str] = {
    "ESTRELA VERMELHA": "1", "ESTRELA": "1",
    "GUAIRACA": "2", "GUAIRACÁ": "2",
    "DEPORTIVO ORIENTAL": "3", "DEPORTIVO": "3",
    "SANKARA": "4",
    "PRIMAVERA": "5",
    "TETO PRETO": "6",
    "PE DE PANO": "7",
    "LOCOMOTIVA": "8",
    "RESISTENCIA ALVIVERDE": "9", "RESISTÊNCIA ALVIVERDE": "9",
    "AZULAO": "10", "AZULÃO": "10",
    "BOLCHESITIO": "11",
    "LINHA ESQUERDA": "12",
    "AMERICA DE CALO": "13", "AMERICA": "13",
    "ATLETICOMUNA": "14",
    "9 DEDOS": "15", "9-DEDOS": "15",
    "RESISTENCIA LATINOFUTURISTA": "16", "LATINOFUTURISTA": "16",
    "LIVERPANCAS": "17", "LIVERPANÇAS": "17",
    "ST PAULO FREIRE": "18", "ST. PAULO FREIRE": "18",
    "COLETIVO SEM FRONTEIRAS": "19", "SEM FRONTEIRAS": "19",
    "TOQUE DE CLASSE": "20",
    "DIAMANTE": "21",
    "DISCAIDA": "22", "DISCAÍDA": "22",
    "BRIGADA LUPICINIA": "23",
    "BAIANOS DE MAUA": "24",
    "CAOS DA VILA": "25", "CAOS DA VILLA": "25",
    "DELAS": "26",
    "FAIXA PRETA": "27",
    "GINGA": "28",
    "MATSUBARA": "29",
    "SAO BENTO": "30", "SÃO BENTO": "30",
    "ATLETICO ZAPATISTA": "31", "ATLETICO ZAPATISTAS": "31", "ZAPATISTA": "31",
    "IV:XX DE NOVEMBRO": "32",
    "AQUI ESTAMOS": "33", "AQUI": "33",
    "IMPERIAL": "34",
    "MATRIARCADO": "35",
    "UMBABARAUMA": "36", "UMBA": "36",
}


def strip_accents(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")


def normalize_team_name(raw: str) -> str:
    return strip_accents(str(raw)).strip().upper()


def resolve_team_id(raw: str) -> str | None:
    key = normalize_team_name(raw)
    return TEAM_ALIASES.get(key)


def normalize_label(raw: str) -> str:
    return strip_accents(str(raw)).strip().upper()


ORDINAL_RE = re.compile(r"(\d+)\s*[ºª°]")


def resolve_competition_slug(round_label: str | None, phase_context: str | None) -> str:
    if round_label:
        norm = normalize_label(round_label)
        for token, slug in COMPETITION_TOKENS:
            if token in norm:
                return slug
    if phase_context:
        norm_phase = normalize_label(phase_context)
        for token, slug in PHASE_CONTEXT_DEFAULT:
            if token in norm_phase:
                return slug
    return DEFAULT_COMPETITION_SLUG


def humanize_round(round_label: str) -> str:
    norm = normalize_label(round_label)
    if "TRIANGULAR" in norm:
        return "Triangular"
    if "QUADRANGULAR" in norm:
        return "Quadrangular"
    if "REPESCAGEM" in norm:
        return "Repescagem"
    if "QUARTAS" in norm:
        return "Quartas de Final"
    if "SEMI" in norm:
        return "Semifinal"
    if "3" in norm and "LUGAR" in norm:
        return "Disputa de 3º Lugar"
    if "FINAL" in norm:
        return "Final"
    match = ORDINAL_RE.search(norm)
    if match and ("DATA" in norm or "RODADA" in norm):
        return f"Rodada {match.group(1)}"
    return round_label.strip().title()


SCORE_PENALTY_RE = re.compile(r"^\s*(\d+)\s*\(\s*(\d+)\s*\)\s*$")


def parse_score(value) -> tuple[int | None, str | None]:
    """Retorna (placar, observação) a partir do valor bruto da célula."""
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        return int(value), None
    text = str(value).strip()
    if not text:
        return None, None
    m = SCORE_PENALTY_RE.match(text)
    if m:
        return int(m.group(1)), f"penaltis:{m.group(2)}"
    if text.upper() == "WO":
        return None, "WO"
    # valor inesperado: preserva como observação e não quebra o pipeline
    return None, f"raw:{text}"


def load_json(name: str) -> dict:
    with open(JSON_DIR / name, encoding="utf-8") as f:
        return json.load(f)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    competitions = load_json("competicoes.json")["competitions"]
    # (season_id, slug) -> competition id / start_date
    comp_lookup: dict[tuple[str, str], dict] = {}
    for c in competitions:
        comp_lookup[(c["season_id"], c["slug"])] = c

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    matches: list[dict] = []
    unmapped_teams: set[str] = set()
    unmapped_competitions: set[str] = set()
    warnings: list[str] = []
    match_counter = 0

    for sheet_name, season_id in SHEET_TO_SEASON.items():
        if sheet_name in SHEETS_TO_SKIP:
            continue
        ws = wb[sheet_name]

        # descobre a última linha com dados nas colunas A-D
        last_row = 0
        for r in range(1, 1001):
            if any(ws.cell(row=r, column=c).value is not None for c in range(1, 5)):
                last_row = r

        current_round_label: str | None = None
        current_phase_context: str | None = None
        # (season_id, slug) -> próxima data disponível para o próximo bloco
        next_date_by_comp: dict[str, date] = {}
        current_block_key: tuple[str, str] | None = None  # (round_label_instance_id, slug)
        block_date_by_key: dict[tuple, date] = {}
        block_counter = 0

        for r in range(3, last_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            d = ws.cell(row=r, column=4).value
            e = ws.cell(row=r, column=5).value
            f = ws.cell(row=r, column=6).value

            if e:
                current_round_label = str(e).strip()
                block_counter += 1
                block_key = (block_counter, current_round_label)
            if f:
                current_phase_context = str(f).strip()

            if not a or not d:
                continue  # linha em branco / separador de bloco

            home_id = resolve_team_id(a)
            away_id = resolve_team_id(d)
            if home_id is None:
                unmapped_teams.add(str(a))
                continue
            if away_id is None:
                unmapped_teams.add(str(d))
                continue

            slug = resolve_competition_slug(current_round_label, current_phase_context)
            comp = comp_lookup.get((season_id, slug))
            if comp is None:
                unmapped_competitions.add(f"{season_id}::{slug} (rodada={current_round_label!r})")
                continue

            comp_id = comp["id"]
            block_key_full = (comp_id, block_counter)
            if block_key_full not in block_date_by_key:
                start = next_date_by_comp.get(comp_id)
                if start is None:
                    start = date.fromisoformat(comp["start_date"])
                block_date_by_key[block_key_full] = start
                next_date_by_comp[comp_id] = start + timedelta(days=7)
            match_date = block_date_by_key[block_key_full]

            score_home, note_home = parse_score(b)
            score_away, note_away = parse_score(c)
            notes_parts = [n for n in (note_home, note_away) if n]
            notes = "; ".join(notes_parts) if notes_parts else None

            match_counter += 1
            match_id = f"match-{match_counter:04d}"
            iso_date = match_date.isoformat()
            matches.append({
                "id": match_id,
                "tournament_id": comp_id,
                "home_team_id": home_id,
                "away_team_id": away_id,
                "score_home": score_home,
                "score_away": score_away,
                "date": iso_date,
                "round": humanize_round(current_round_label) if current_round_label else None,
                "location": None,
                "status": "finished",
                "notes": notes,
                "created_at": f"{iso_date}T00:00:00Z",
                "updated_at": f"{iso_date}T00:00:00Z",
            })

    print(f"Total de partidas geradas: {len(matches)}")
    by_season: dict[str, int] = {}
    for m in matches:
        comp = next(c for c in competitions if c["id"] == m["tournament_id"])
        by_season[comp["season_id"]] = by_season.get(comp["season_id"], 0) + 1
    for season_id, count in by_season.items():
        print(f"  {season_id}: {count} partidas")

    if unmapped_teams:
        print("\n[ALERTA] Times não mapeados (partidas ignoradas):")
        for t in sorted(unmapped_teams):
            print("  -", t)
    if unmapped_competitions:
        print("\n[ALERTA] Competições não encontradas (partidas ignoradas):")
        for t in sorted(unmapped_competitions):
            print("  -", t)
    if warnings:
        print("\n[AVISOS]")
        for w in warnings:
            print("  -", w)

    if args.dry_run:
        print("\n--dry-run: nenhum arquivo foi escrito.")
        return

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump({"matches": matches}, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"\nArquivo escrito em: {OUT_PATH}")


if __name__ == "__main__":
    main()
